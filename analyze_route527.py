# -*- coding: utf-8 -*-
"""Per-stop co-stopping analysis for route 527 (Baiyun Station - Shixi) + competitiveness.

Output is direction-aware: one row per 527 direction x stop x other route direction.
"""
import json
import math
import os
import re
import struct

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "brt_output"
MAIN = "527路"


def route_num(route_cn):
    m = re.match(r"([^（(]+)", route_cn)
    return m.group(1).strip() if m else route_cn


def od_of(route_cn):
    m = re.search(r"\((.*)\)\s*$", route_cn)
    return m.group(1) if m else ""


def read_shp_lines(path):
    with open(path, "rb") as f:
        data = f.read()
    shapes = []
    off = 100
    while off + 8 <= len(data):
        rec_len = struct.unpack(">i", data[off + 4:off + 8])[0] * 2
        if rec_len <= 0 or off + 8 + rec_len > len(data):
            break
        content = data[off + 8:off + 8 + rec_len]
        if len(content) >= 4:
            st = struct.unpack("<i", content[0:4])[0]
            if st == 3:
                nparts, npts = struct.unpack("<ii", content[36:44])
                parts = struct.unpack("<" + "i" * nparts, content[44:44 + 4 * nparts])
                raw = struct.unpack("<" + "d" * (2 * npts),
                                    content[44 + 4 * nparts:44 + 4 * nparts + 16 * npts])
                shapes.append([(raw[i], raw[i + 1]) for i in range(0, 2 * npts, 2)])
            else:
                shapes.append([])
        off += 8 + rec_len
    return shapes


def read_shp_points(path):
    with open(path, "rb") as f:
        data = f.read()
    pts = []
    off = 100
    while off + 8 <= len(data):
        rec_len = struct.unpack(">i", data[off + 4:off + 8])[0] * 2
        if rec_len <= 0 or off + 8 + rec_len > len(data):
            break
        content = data[off + 8:off + 8 + rec_len]
        if len(content) >= 4:
            st = struct.unpack("<i", content[0:4])[0]
            if st == 1:
                x, y = struct.unpack("<dd", content[4:20])
                pts.append((x, y))
            else:
                pts.append(None)
        off += 8 + rec_len
    return pts


def bearing(a, b):
    if a is None or b is None or a == b:
        return None
    dy = b[1] - a[1]
    dx = (b[0] - a[0]) * math.cos(math.radians((a[1] + b[1]) / 2))
    return math.degrees(math.atan2(dx, dy)) % 360


def angle_diff(a, b):
    if a is None or b is None:
        return None
    d = abs(a - b) % 360
    return min(d, 360 - d)


def poly_bearing(rc, spt, routes_idx, line_shapes):
    i = routes_idx.get(rc)
    if i is None or spt is None:
        return None
    shape = line_shapes[i]
    if len(shape) < 2:
        return None
    best = min(range(len(shape)),
               key=lambda k: (shape[k][0] - spt[0]) ** 2 + (shape[k][1] - spt[1]) ** 2)
    lo, hi = max(best - 2, 0), min(best + 2, len(shape) - 1)
    if lo >= hi:
        return None
    return bearing(shape[lo], shape[hi])


def main():
    os.makedirs(OUT, exist_ok=True)
    with open("_parsed_bus.json", "r", encoding="utf-8") as f:
        D = json.load(f)
    routes = pd.DataFrame(D["routes"])
    stops = pd.DataFrame(D["stops"])
    su = pd.DataFrame(D["stops_unique"])

    r527 = routes[routes["route_cn"].str.startswith(MAIN, na=False)]
    routes_idx = {rc: i for i, rc in enumerate(routes["route_cn"])}
    line_shapes = read_shp_lines("guangzhou_bus_routes.shp")
    stop_pts = read_shp_points("guangzhou_bus_stops.shp")
    print("527 records in routes table:", len(r527))
    print(r527[["route_cn", "distance", "total_stop", "basic_prc", "company_cn"]].to_string())

    s527 = stops[stops["route_cn"].str.startswith(MAIN, na=False)].copy()
    s527_main = s527[~s527["route_cn"].str.startswith("527路机电", na=False)].copy()
    dir_names = sorted(s527_main["route_cn"].unique())
    print("\n527 main directions:", dir_names)

    # stop -> other routes map (exclude all 527 records incl. shuttle), direction-specific
    excl = stops["route_cn"].str.startswith("527", na=False)
    other = stops[~excl]
    stop2routes = {sid: sorted(g["route_cn"].unique().tolist())
                   for sid, g in other.groupby("stop_id")}

    # route info: number -> direction-specific ODs + length/stops/price/company
    rinfo = {}
    for _, row in routes.iterrows():
        n = route_num(row["route_cn"])
        if n not in rinfo:
            rinfo[n] = {"ods": set(), "len": row["distance"], "stops": row["total_stop"],
                        "price": row["basic_prc"], "company": row["company_cn"] or "",
                        "variants": set()}
        rinfo[n]["ods"].add(od_of(row["route_cn"]))
        rinfo[n]["variants"].add(row["route_cn"])

    # ---------- detail: one row per (527 direction, stop, other route number) ----------
    # Within one 527 direction x stop, each other line appears once. When the line
    # has records for both directions at the same stop_id, the travel direction is
    # inferred from route geometry (bearing of the route polyline at the stop).
    # Only if a clear match exists is a single direction kept; otherwise both are
    # kept and marked as undetermined.
    det_rows = []
    for rc_dir in dir_names:
        g = s527_main[s527_main["route_cn"] == rc_dir].sort_values("sequence")
        for _, r in g.iterrows():
            spt = stop_pts[r.name]
            b527v = poly_bearing(rc_dir, spt, routes_idx, line_shapes)
            by_num = {}
            for orc in stop2routes.get(r["stop_id"], []):
                by_num.setdefault(route_num(orc), set()).add(orc)
            for n in sorted(by_num):
                variants = sorted(by_num[n])
                if len(variants) == 1:
                    orc = variants[0]
                    judge = "单方向"
                    line_col = orc
                    od_col = od_of(orc)
                else:
                    scored = []
                    for orc in variants:
                        b = poly_bearing(orc, spt, routes_idx, line_shapes)
                        d = angle_diff(b, b527v)
                        scored.append((d if d is not None else 999, orc))
                    scored.sort()
                    best_d, best_orc = scored[0]
                    second_d = scored[1][0] if len(scored) > 1 else None
                    margin = second_d - best_d if second_d is not None else None
                    if margin is not None and margin >= 45:
                        judge = "按走向匹配"
                        line_col = best_orc
                        od_col = od_of(best_orc)
                    else:
                        judge = "双向待定（走向无法区分）"
                        line_col = n
                        od_col = " / ".join(sorted({od_of(v) for v in variants}))
                det_rows.append({
                    "527方向": rc_dir,
                    "序号": int(r["sequence"]),
                    "站名": r["name_cn"],
                    "站点ID": r["stop_id"],
                    "共线线路": line_col,
                    "起讫站": od_col,
                    "线路里程": rinfo[n]["len"],
                    "票价": rinfo[n]["price"],
                    "运营公司": rinfo[n]["company"],
                    "方向判定": judge,
                })
    det = pd.DataFrame(det_rows).drop_duplicates()
    print("\ndetail rows:", len(det), "(unique)")
    det.to_csv(os.path.join(OUT, "527_共线明细.csv"), index=False, encoding="utf-8-sig")

    # ---------- overview: one row per (527 direction, stop) ----------
    ov_rows = []
    for rc_dir in dir_names:
        g = s527_main[s527_main["route_cn"] == rc_dir].sort_values("sequence")
        for _, r in g.iterrows():
            orcs = stop2routes.get(r["stop_id"], [])
            nums = sorted(set(route_num(x) for x in orcs))
            ov_rows.append({
                "527方向": rc_dir,
                "序号": int(r["sequence"]),
                "站名": r["name_cn"],
                "站点ID": r["stop_id"],
                "其他线路数": len(nums),
                "方向记录数": len(orcs),
                "共线线路": "、".join(nums),
            })
    ov = pd.DataFrame(ov_rows)
    ov.to_csv(os.path.join(OUT, "527_每站共线概览.csv"), index=False, encoding="utf-8-sig")
    print("overview rows:", len(ov))

    # ---------- ranking: network-level with direction breakdown ----------
    stop_ids_dir = {}
    for rc_dir in dir_names:
        stop_ids_dir[rc_dir] = set(s527_main[s527_main["route_cn"] == rc_dir]["stop_id"])
    all_ids = set().union(*stop_ids_dir.values())

    overlap = {}
    for sid in all_ids:
        for rc in stop2routes.get(sid, []):
            n = route_num(rc)
            overlap.setdefault(n, set()).add(sid)
    score_rows = []
    for n, sids in overlap.items():
        score_rows.append({
            "线路": n,
            "共线站点数": len(sids),
            "去程共线站数": len(sids & stop_ids_dir[dir_names[0]]),
            "回程共线站数": len(sids & stop_ids_dir[dir_names[1]]),
            "覆盖527站点比例": 0,
            "起讫站": " / ".join(sorted(rinfo[n]["ods"])),
            "线路里程": rinfo[n]["len"],
            "站点数": rinfo[n]["stops"],
            "票价": rinfo[n]["price"],
            "运营公司": rinfo[n]["company"],
        })
    score = pd.DataFrame(score_rows)
    score["共线站点数"] = score["共线站点数"].astype(int)
    score["去程共线站数"] = score["去程共线站数"].astype(int)
    score["回程共线站数"] = score["回程共线站数"].astype(int)
    score["覆盖527站点比例"] = score["共线站点数"] / len(all_ids)
    score = score.sort_values(["共线站点数", "去程共线站数", "回程共线站数"],
                              ascending=False).reset_index(drop=True)
    print("\n=== 与527主线共线 Top30（含方向拆分）===")
    print(score.head(30).to_string())
    print("共线线路总数:", len(score))
    score.to_csv(os.path.join(OUT, "527_共线线路排行.csv"), index=False, encoding="utf-8-sig")

    # ---------- excel ----------
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F77B4")
    hdr_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, df):
        ws.append(list(df.columns))
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in df.iterrows():
            ws.append(list(row))
        for c in range(1, len(df.columns) + 1):
            width = max(len(str(df.columns[c - 1])),
                        max((len(str(v)) for v in df.iloc[:, c - 1]), default=1))
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), 60)
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "每站共线明细"
    write_sheet(ws1, det)
    ws2 = wb.create_sheet("每站共线概览")
    write_sheet(ws2, ov)
    ws3 = wb.create_sheet("共线线路排行")
    write_sheet(ws3, score)
    wb.save(os.path.join(OUT, "527_共线分析.xlsx"))
    print("excel saved")

    # ---------- json for report ----------
    with open(os.path.join(OUT, "527_data.json"), "w", encoding="utf-8") as f:
        json.dump({
            "main": r527[["route_cn", "distance", "total_stop", "basic_prc", "company_cn"]].to_dict("records"),
            "dir_names": dir_names,
            "overview": ov.to_dict("records"),
            "detail": det.to_dict("records"),
            "score": score.to_dict("records"),
            "rinfo": {k: {**v, "ods": sorted(v["ods"]), "variants": sorted(v["variants"])}
                      for k, v in rinfo.items()},
            "n_unique_stops": len(all_ids),
        }, f, ensure_ascii=False, indent=1)
    print("json saved")


if __name__ == "__main__":
    main()

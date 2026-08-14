# -*- coding: utf-8 -*-
"""Build BRT analysis report (Markdown) and Excel workbook."""
import json
import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "brt_output"


def base(line):
    m = re.match(r"B(\d+)", line.upper())
    return "B" + m.group(1) if m else line


def main():
    os.makedirs(OUT, exist_ok=True)
    with open("_parsed_bus.json", "r", encoding="utf-8") as f:
        D = json.load(f)
    routes = pd.DataFrame(D["routes"])
    stops = pd.DataFrame(D["stops"])
    su = pd.DataFrame(D["stops_unique"])
    segments = pd.DataFrame(D["segments"])

    mask = routes["route_cn"].str.upper().str.startswith("B", na=False)
    br = routes[mask].copy().reset_index(drop=True)
    br["base"] = br["route_cn"].apply(base)
    bnames = set(br["route_cn"])
    bs = stops[stops["route_cn"].isin(bnames)]
    bseg = segments[segments["s_stopid"].isin(bs["stop_id"]) | segments["e_stopid"].isin(bs["stop_id"])]

    # time stats
    st = br.dropna(subset=["start_time"]).copy()
    st["hhmm"] = st["start_time"].astype(int)
    et = br.dropna(subset=["end_time"]).copy()
    et["hhmm"] = et["end_time"].astype(int)
    start_med = f'{int(st["hhmm"].median() // 100):02d}:{int(st["hhmm"].median() % 100):02d}'
    end_med = f'{int(et["hhmm"].median() // 100):02d}:{int(et["hhmm"].median() % 100):02d}'

    # corridor counts
    corridor = ["棠下村", "棠东", "华景新城", "车陂", "师大暨大", "黄村", "天朗明居",
                "体育中心", "上社", "石牌桥", "岗顶", "珠村", "茅岗", "莲溪"]
    corr = [(n, int(bs[bs["name_cn"] == n]["route_cn"].nunique())) for n in corridor]
    corr_sorted = sorted(corr, key=lambda t: -t[1])

    # company
    comp = br["company_cn"].value_counts(dropna=False)
    n_company = br["company_cn"].notna().sum()

    bases = sorted(br["base"].unique(), key=lambda s: int(s[1:]))
    g = br.groupby("base").agg(
        cnt=("route_cn", "count"),
        mn=("distance", "min"), med=("distance", "median"), mx=("distance", "max"),
        stop_min=("total_stop", "min"), stop_max=("total_stop", "max"),
    ).reindex(bases)

    rows = []
    for _, r in br.sort_values(["base", "distance"]).iterrows():
        rows.append({
            "线路": r["route_cn"], "线路号": r["base"],
            "里程(km)": round(float(r["distance"]), 2), "站点数": int(r["total_stop"]),
            "平均站距(m)": round(float(r["distance"]) * 1000 / max(int(r["total_stop"]) - 1, 1), 1),
            "首班": r["start_time"], "末班": r["end_time"],
            "运营公司": r["company_cn"] or "", "起点": r["s_stop_cn"], "终点": r["e_stop_cn"],
        })
    detail = pd.DataFrame(rows)

    # stop-level detail
    stop_rows = []
    for _, r in bs.sort_values(["route_cn", "sequence"]).iterrows():
        stop_rows.append({
            "线路": r["route_cn"], "序号": int(r["sequence"]),
            "站名": r["name_cn"], "站点ID": r["stop_id"],
        })
    stop_detail = pd.DataFrame(stop_rows)

    # corridor stations full info
    corr_rows = []
    for n, c in corr:
        info = su[su["stop_cn"] == n]
        corr_rows.append({
            "站点": n, "B线数": c,
            "全部线路覆盖数": int(info["num"].max()) if len(info) else None,
        })
    corr_df = pd.DataFrame(corr_rows)

    # ---------- Excel ----------
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F77B4")
    hdr_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, df):
        ws.append(list(df.columns))
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in df.iterrows():
            ws.append(list(row))
        for c in range(1, len(df.columns) + 1):
            width = max(len(str(df.columns[c - 1])), max((len(str(v)) for v in df.iloc[:, c - 1]), default=1))
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), 60)
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "B线路明细"
    write_sheet(ws1, detail)
    ws2 = wb.create_sheet("B线路站点明细")
    write_sheet(ws2, stop_detail)
    ws3 = wb.create_sheet("BRT走廊站点覆盖")
    write_sheet(ws3, corr_df)
    ws4 = wb.create_sheet("线路号汇总")
    sum_df = g.reset_index().rename(columns={
        "index": "线路号", "cnt": "线路数", "mn": "最短里程", "med": "中位里程",
        "mx": "最长里程", "stop_min": "最少站点", "stop_max": "最多站点"})
    write_sheet(ws4, sum_df)
    wb.save(os.path.join(OUT, "BRT线路明细.xlsx"))
    print("excel saved")

    # ---------- Markdown report ----------
    md = []
    md.append("# 广州 BRT（B 字头线路）数据分析报告\n")
    md.append("## 1. 数据概况\n")
    md.append(f"- B 字头线路共 **{len(br)} 条**（含往返方向和快线/分线变体），归属 **{len(bases)} 个基础线路号**（B1–B31）。")
    md.append(f"- 全部为广州市内线路，状态正常（status=1），均非循环线；票价统一 **2 元**（基本票价=全程票价）。")
    md.append(f"- B 线站点记录 **{len(bs)} 条**，去重后 **{bs['stop_id'].nunique()} 个站点**；")
    md.append(f"- 线路长度中位数 **{br['distance'].median():.1f} km**（全市中位数 17.1 km），站点数中位数 **{int(br['total_stop'].median())} 站**（全市 23 站）。")
    md.append("\n## 2. 里程与站点规模\n")
    md.append("| 指标 | B 线 | 全市全部线路 |")
    md.append("|---|---|---|")
    md.append(f"| 线路数 | {len(br)} | {len(routes)} |")
    md.append(f"| 平均里程 (km) | {br['distance'].mean():.1f} | {routes['distance'].mean():.1f} |")
    md.append(f"| 里程中位数 (km) | {br['distance'].median():.1f} | {routes['distance'].median():.1f} |")
    md.append(f"| 里程范围 (km) | {br['distance'].min():.1f} – {br['distance'].max():.1f} | {routes['distance'].min():.1f} – {routes['distance'].max():.1f} |")
    md.append(f"| 平均站点数 | {br['total_stop'].mean():.1f} | {routes['total_stop'].mean():.1f} |")
    md.append(f"| 站点数范围 | {int(br['total_stop'].min())} – {int(br['total_stop'].max())} | {int(routes['total_stop'].min())} – {int(routes['total_stop'].max())} |")
    md.append(f"| 平均站距 (km) | {((br['distance'] / (br['total_stop'] - 1))).mean():.2f} | {((routes['distance'] / (routes['total_stop'] - 1))).mean():.2f} |")
    md.append("\n- 最短：**B14路**(棠德花苑总站–体育中心)，约 10.0 km / 10 站；")
    md.append("- 最长：**B18路快线**(永泰路口–汇彩路北总站)，约 28.7 km；B18路 与 B21路、B24路、B6路等也超过 27 km。")
    md.append("- 基础线路号中 B4 变体最多（B4、B4A、B4B 共 6 条记录），B1/B2/B6/B7/B18 各有 4 条（含快线/短线）。")
    md.append("\n## 3. BRT 走廊站点覆盖（中山大道一带）\n")
    md.append("下表为停靠各走廊站点的 B 线数量（按站名统计）：\n")
    md.append("| 站点 | B 线数 | 站点 | B 线数 |")
    md.append("|---|---|---|---|")
    for i in range(0, len(corr_sorted), 2):
        a = corr_sorted[i]
        b = corr_sorted[i + 1] if i + 1 < len(corr_sorted) else ("", "")
        md.append(f"| {a[0]} | {a[1]} | {b[0]} | {b[1]} |")
    md.append("\n- 核心段（华景新城—车陂—棠东—棠下村）被 **25–35 条** B 线共用，是真正的 BRT 主干走廊；")
    md.append("- 向东（珠村 16 条、茅岗 12 条、莲溪 4 条）线路明显减少，BRT 专用道东段利用度较低；")
    md.append("- 站点数据中还包含 BRT 子站（如 BRT岗顶N1、BRT石牌桥S2、BRT车陂S1 等），说明同一走廊站按方向/子站拆分为多个站点 ID。")
    md.append("\n## 4. 主要枢纽与起终点\n")
    md.append("B 线起终点出现频次最高的枢纽：**车陂总站（4 条）**、体育中心/体育中心总站（5 条）、汇彩路总站/汇彩路北总站（6 条）、棠德花苑总站（3 条）、广州火车东站总站（3 条）、广州火车站总站（2 条）、同和路总站（2 条）等，整体呈“天河核心区—黄埔—白云（同和/永泰）—海珠（宝岗/海珠客运站）”放射状。")
    md.append("\n## 5. 运营公司\n")
    md.append(f"共涉及 {n_company} 家明确的运营主体，前三：")
    for name, c in comp.head(3).items():
        md.append(f"- {name}：{int(c)} 条")
    md.append(f"\n其余线路由巴士集团属下车队、一汽巴士、顺途等运营。")
    md.append("\n## 6. 运营时间\n")
    md.append(f"- 首班时间有记录的 {len(st)} 条，中位数 **{start_med}**，最常见 06:00 与 06:30；")
    md.append(f"- 末班时间有记录的 {len(et)} 条，中位数 **{end_med}**，最常见 22:30 与 22:00。")
    md.append("\n## 7. 小结\n")
    md.append("1. B 线是一个规模 69 条（含变体）、平均 21 km、平均 26 站的快速公交网络，票价统一 2 元，整体比全市普通线路更长、站点更多。")
    md.append("2. 走廊高度集中在中山大道 BRT 核心段（棠下村、棠东、车陂、华景新城一带），呈现明显的东西向主干 + 南北向支线的放射结构。")
    md.append("3. B18/B21/B24/B31 等是网络中的长距离骨干，B14 是最短的区间短线。")
    md.append("4. 数据存在少量子站拆分与 2 条线路站点数与线路属性略有出入（B30、B6快线），分析时建议按站点 ID + 站名双重核对。")
    md.append("\n---\n")
    md.append("报告由脚本自动生成，明细见 `BRT线路明细.xlsx`，图表见 `brt_map.png` / `brt_lengths.png` / `brt_stops.png` / `brt_corridor.png`。")

    with open(os.path.join(OUT, "BRT分析报告.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    print("report saved")


if __name__ == "__main__":
    main()
